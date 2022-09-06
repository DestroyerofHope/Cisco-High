# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

class CH_Automation():
    
    def __init__(self):
        self.uniquesites = {}
        self.xlsxs = {}
    
    def addXLSX(self, xlsx, name):
        '''Created a workbook to manipulate the xlsx file; has to be fill address in this format:''' #var.addXLSX(r"C:\Users\jasoli3\Downloads\2022 Extern Details.xlsx")
        workbook = load_workbook(filename=xlsx)
        self.xlsxs[name] = [name + '1', name + '2']
        self.xlsxs[name][0] = workbook.active
        self.xlsxs[name][1] = pd.read_excel(xlsx)
        self.xlsxs[name][1].dropna(how = 'all', inplace = True)

    def sort(self, name, columnnum, namecolumn = None):
        columnname = self.xlsxs[name][0].cell(row = 1, column = columnnum).value
        if namecolumn == None:
            self.xlsxs[name][1].sort_values(by=columnname, inplace = True)
        if type(namecolumn) == int:
            columnname1 = self.xlsxs[name][0].cell(row = 1, column = namecolumn).value
            self.xlsxs[name][1].sort_values(by=[columnname, columnname1], inplace = True)
        self.xlsxs[name][1].reset_index(drop = True, inplace = True)
        return self.xlsxs[name][1]

    def columnrange(self, name, column, namecolumn = None):
        '''uniquesites works if the Excel self.xlsxs[name][0] is sorted by site and the column of site is names 'Sites'.//
        initialrow and finalrow are type int for the initial and final roself.xlsxs[name][0] to search through. The function returns//
        a dict of the unique sites in the Excel self.xlsxs[name][0] and the range of the roself.xlsxs[name][0] it spans.'''
        self.uniquesites = {}
        self.sort(name, column, namecolumn)
        for i in self.xlsxs[name][1].index:
            if i == 0:
                self.uniquesites[self.xlsxs[name][1].values[i][column-1]] = [i]
            else:
                if i < self.xlsxs[name][1].index[len(self.xlsxs[name][1].index)-1]:
                    if self.xlsxs[name][1].values[i][column-1] != self.xlsxs[name][1].values[i+1][column-1]:
                        self.uniquesites[self.xlsxs[name][1].values[i][column-1]].append(i)
                        self.uniquesites[self.xlsxs[name][1].values[i+1][column-1]] = [i+1]
                else:
                    self.uniquesites[self.xlsxs[name][1].values[i][column-1]].append(i)
        return self.uniquesites
    
    def uniquevalcount(self, name, statcolumn, rows = None):
        uniqval = {}
        columnname = self.xlsxs[name][0].cell(row = 1, column = statcolumn).value
        vals = (self.xlsxs[name][1][columnname].dropna(how = 'all')).to_numpy()
        if type(rows) == list:
            vals = vals[rows[0]:rows[1]]
        uniqvals = np.unique(vals)
        for i in uniqvals:
            uniqval[i] = 0
        for i in vals:
            for j in uniqval:
                if i == j:
                    uniqval[j] += 1
        return uniqval

class Extern_Details(CH_Automation):
    
    def __init__(self, xlsx):
        CH_Automation.__init__(self)
        CH_Automation.addXLSX(self, xlsx, 'Extern_Details')
        self.uniquevals = {}
        self.name = 'Extern_Details'
        
        
class Demographics(Extern_Details):
    
    def __init__(self, xlsx):
        Extern_Details.__init__(self, xlsx)
        self.demostat = {}
    
    def demostatistics(self, statcolumn, rows = None):
            vals, tot = CH_Automation.uniquevalcount(self, self.name, statcolumn, rows), 0
            for i in vals:
                tot += vals[i]
            for i in vals:
                vals[i] = str(round(vals[i]/tot * 100, 2)) + ' %'
            return vals
    
    def sitestat(self, sitename, sitecolumn, statcolumn):
        sites, siterange = CH_Automation.columnrange(self, self.name, sitecolumn), ''
        for i in sites:
            if i == sitename:
                siterange = sites[i]
        sitestat = self.demostatistics(statcolumn, siterange)
        return sitestat
    
    def piechart(self, statcolumn):
        title, label = self.sheet.cell(row = 1, column = statcolumn).value, []
        uniqval = CH_Automation.uniquevalcount(self, self.name, statcolumn)
        for i in uniqval:
            label.append(i)
        values = []
        for i in uniqval:
            values.append(uniqval[i])
        values = np.array(values)
        plt.pie(values, labels = label)
        plt.title(title)
        plt.show()
    
    def sitepiechart(self, sitename, sitecolumn, statcolumn):
        title, label = self.xlsxs[self.name][0].cell(row = 1, column = statcolumn).value, []
        sites, siterange = CH_Automation.columnrange(self, self.name, sitecolumn), ''
        for i in sites:
            if i == sitename:
                siterange = sites[i]
        uniqval = CH_Automation.uniquevalcount(self, self.name, statcolumn, siterange)
        for i in uniqval:
            label.append(i)
        values = []
        for i in uniqval:
            values.append(uniqval[i])
        values = np.array(values)
        plt.pie(values, labels = label)
        plt.title(title)
        plt.show()    
    
class Capstone_Groups(Extern_Details):
    
    def __init__(self, xlsx):
        Extern_Details.__init__(self, xlsx)
        self.attr = {}
        self.groups = []
        self.groupnum = 0
        self.groupsize = []
        self.groupavg = []
        self.placegroup = []
    
    def assignNum(self, attribute, val = 'Preset'):
        '''This function assigns the numerical value associate with a certian demographic attribute. attribute is a str. Outputs an int'''
        if val == 'Preset':
            if attribute == 'RTP':
                return 0
            elif attribute == 'Richardson':
                return .33
            elif attribute == 'Herndon':
                return .67
            elif attribute == 'California':
                return 1
            elif attribute == 'Atlanta':
                return 0
            elif attribute == 'Chicago':
                return .25
            elif attribute == 'NYC':
                return .5
            elif attribute == 'Toronto':
                return .75
            elif attribute == 'St. Louis':
                return 1
            elif attribute == 'Male' or attribute == 'Man':
                return 0
            elif attribute == 'Female' or attribute == 'Woman':
                return 1
            elif attribute == 'Transgender':
                return 1.5
            elif attribute in ('African American', 'Black', 'Black / African American', '/African American'):
                return 0
            elif attribute == 'Latin / Spanish' or attribute == 'Spanish / Hispanic / Latino':
                return .17
            elif attribute == 'Pacific Islander':
                return .33
            elif attribute == 'Caucasian' or attribute == 'White / Caucasian':
                return .5
            elif attribute == 'Asian':
                return .67
            elif attribute == 'Other':
                return .83
            elif attribute == 'Prefer Not to Answer':
                return 1
            elif attribute == None or pd.isna(attribute):
                return .5
            
            elif type(val) == dict:
                for i in val:
                    if attribute == i:
                        return val[i]
                    elif attribute == None or pd.isna(attribute):
                        return .5 
        
    def defAttributes(self, namecolumn, sitecolumn, gendercolumn, racecolumn, rows = None):
        '''defAttribute creates a dict with the names of the externs and the numerical value associated with them. initialrow & finalrow are ints, representing the initial and final
            rows the algorithm searches through, namecolumn, sitecolumn, gendercolumn, and racecolumn are ints, they are the columns those details are located in.'''
        dataframe = self.xlsxs[self.name][1].values
        if rows == None:
            for i in range(len(dataframe)):
                self.attr[dataframe[i][namecolumn-1]] = [self.assignNum(dataframe[i][sitecolumn-1]), self.assignNum(dataframe[i][gendercolumn-1]), self.assignNum(dataframe[i][racecolumn-1])]
        elif type(rows) == list:
            for i in range(rows[0], rows[1]):
                self.attr[dataframe[i][namecolumn-1]] = [self.assignNum(dataframe[i][sitecolumn-1]), self.assignNum(dataframe[i][gendercolumn-1]), self.assignNum(dataframe[i][racecolumn-1])]
        return self.attr
        
    def groupSizes(self, groupnum):
        '''groupSizes splits the total number of externs into equal groups, groupnum is an int representing the number of groups the externs are to be split into.
            Outputs a list of group sizes'''
        self.groupnum = groupnum
        num, remainder = np.floor(
            len(self.attr)/self.groupnum), (len(self.attr) % self.groupnum)
        for i in range(self.groupnum):
            self.groupsize.append(num)
        for i in range(remainder):
            self.groupsize[i] += 1
        for i in range(len(self.groupsize)):
            self.groupsize[i] = int(self.groupsize[i])
        return self.groupsize

    def assignTarget(self):
        '''assignTarget calculates the target numbers for different demographic attributes based on the demographic ratios in the sample. Outputs a list of target values.'''
        targets = []
        for i in range(len(list(self.attr.values())[0])):
            target = 0
            for j in range(len(self.attr)):
                target += list(self.attr.values())[j][i]
            target = target/int(len(self.attr))
            targets.append(target)
        return targets

    def createGroup(self, groups, iterations, groupiter):
        '''createGroup recursively creates a group based on the numerical values calculated in assignTarget. iterations and groupiter are ints, groups is a dict
            iterations is the number of iterations the function has done already and groupiter is the number of people in the created group, 
            groups is the dict of externs and demographic values. Outputs a list of capstone groups and the experimental demographic numerical values for the groups.'''
        targets = self.assignTarget()
        if iterations == 0:
            self.placegroup.append(
                [list(groups.keys())[0], list(groups.values())[0]])
            del groups[list(groups.keys())[0]]
            self.groupavg.append(
                [list(self.placegroup)[0][1][0], len(self.placegroup)])
            self.groupavg.append(
                [list(self.placegroup)[0][1][1], len(self.placegroup)])
            self.groupavg.append(
                [list(self.placegroup)[0][1][2], len(self.placegroup)])
        avgnum = {}
        for name in groups:
            avgs = [(list(groups[name])[0] + (self.groupavg[0][0]*self.groupavg[0][1]))/(self.groupavg[0][1]+1), (list(groups[name])[1] + (self.groupavg[1]
                    [0]*self.groupavg[1][1]))/(self.groupavg[1][1]+1), (list(groups[name])[2] + (self.groupavg[2][0]*self.groupavg[2][1]))/(self.groupavg[2][1]+1)]
            for j in range(len(targets)):
                avgs[j] = abs(targets[j] - avgs[j])
            placeavgnum = sum(avgs)/3
            avgnum[name] = placeavgnum
        choice = list(avgnum.keys())[0]
        for name in avgnum:
            if avgnum[name] < avgnum[choice]:
                choice = name
        self.placegroup.append([choice, groups[choice]])
        self.groupavg[0][0] = self.groupavg[0][0] * \
            self.groupavg[0][1] + groups[choice][0]
        self.groupavg[1][0] = self.groupavg[1][0] * \
            self.groupavg[1][1] + groups[choice][1]
        self.groupavg[2][0] = self.groupavg[2][0] * \
            self.groupavg[2][1] + groups[choice][2]
        self.groupavg[0][1] += 1
        self.groupavg[1][1] += 1
        self.groupavg[2][1] += 1
        self.groupavg[0][0] = self.groupavg[0][0]/self.groupavg[0][1]
        self.groupavg[1][0] = self.groupavg[1][0]/self.groupavg[1][1]
        self.groupavg[2][0] = self.groupavg[2][0]/self.groupavg[2][1]
        groups.pop(choice)
        choice = ''
        if iterations < groupiter - 2:
            self.createGroup(groups, iterations + 1, groupiter)
        if iterations <= groupiter - 2:
            del groups
            return self.placegroup, self.groupavg

    def assignGroups(self):
        '''Calls createGroup multiple times to create the appropriate number of groups for the externs. It returns a list of the groups.'''
        groupiter, self.placegroup, grouptarget = self.groupsize, [], []
        groups = self.attr.copy()
        for size in groupiter:
            self.groupavg, self.placegroup = [], []
            groupnames = []
            group, groupavg = self.createGroup(groups, 0, size)
            grouptarget.append([groupavg[0][0], groupavg[1][0], groupavg[2][0]])
            for i in range(len(group)):
                groupnames.append(group[i][0])
            self.groups.append(groupnames)
        return self.groups, grouptarget
    
    def targetError(self):
        '''targetError calculates the percent error of the groups made by assignGroups in comparison to the calculated target numbers in assingTarget
            It returns a list of the average percent errors of each group.'''
        target = self.assignTarget()
        groups, grouptarget = self.assignGroups()
        error = []
        for i in range(len(grouptarget)):
            error1 = abs(grouptarget[i][0] - target[0])/ target[0]
            error2 = abs(grouptarget[i][1] - target[1])/ target[1]
            error3 = abs(grouptarget[i][2] - target[2])/ target[2]
            error.append(str(round(((error1 + error2 + error3)/3)*100, 2)) + ' %')
        return error

class CH_Swag(Extern_Details):
    
    def __init__(self, xlsx):
        pass

class Daily_Survey(CH_Automation):
    
    def __init__(self, xlsx):
        CH_Automation.__init__(self)
        CH_Automation.addXLSX(self, xlsx, 'Daily_Survey')
        self.uniquevals = {}
        self.sessionrating = {}
        self.name = 'Daily_Survey'
        self.sheet = self.xlsxs[self.name][0]
        self.df = self.xlsxs[self.name][1]
        
    def sessionratings(self, sessioncolumn, ratingcolumn, rows = None, sort = False):
        if type(rows) == list and sort == True:
            self.sort('Daily_Survey', 2)
            sessionname, ratingname = self.sheet.cell(row = 1, column = sessioncolumn).value, self.sheet.cell(row = 1, column = ratingcolumn).value
            sessionval, ratingval = self.df[sessionname].to_numpy()[rows[0]:rows[1]], self.df[ratingname].to_numpy()[rows[0]:rows[1]]
            sessionnum = CH_Automation.uniquevalcount(self, self.name, sessioncolumn, rows)
            uniqsessions = {}
            uniq = np.unique((self.df[sessionname].dropna(how = 'all')).to_numpy()[rows[0]:rows[1]])
        else:
            sessionname, ratingname = self.sheet.cell(row = 1, column = sessioncolumn).value, self.sheet.cell(row = 1, column = ratingcolumn).value
            sessionval, ratingval = self.df[sessionname].to_numpy(), self.df[ratingname].to_numpy()
            sessionnum = CH_Automation.uniquevalcount(self, self.name, sessioncolumn, rows)
            uniqsessions = {}
            uniq = np.unique(((self.df[sessionname].dropna(how = 'all')).to_numpy())[rows[0]:rows[1]])
        #print(sessionval, ratingval)
        print(sessionnum)
        for session in uniq:
            uniqsessions[session] = 0
        for i in range(len(sessionval)):
            for session in uniqsessions:
                if sessionval[i] == session:
                    if pd.notna(ratingval[i]):
                        uniqsessions[session] += int(str(ratingval[i])[0])
        #return uniqsessions, sessionnum
    
    def totalsessionratings(self, sessioncolumn, ratingcolumn, rows = None, sort = None):
        sessionnum = {}
        if len(sessioncolumn) == len(ratingcolumn) and list in (type(sessioncolumn), type(ratingcolumn)):    
            for i in range(len(sessioncolumn)):
                if i == 0:
                    ratingnum = self.sessionratings(sessioncolumn[0], ratingcolumn[0], rows, sort)[0]
                else:
                    ratingnum.update(self.sessionratings(sessioncolumn[i], ratingcolumn[i], rows, sort)[0])
            for session in ratingnum:
                ratingnum[session] = 0
                sessionnum[session] = 0
            for i in range(len(sessioncolumn)):
                rating, sessions = self.sessionratings(sessioncolumn[i], ratingcolumn[i], rows, sort)
                for session in ratingnum:
                    try:
                        ratingnum[session] += int(rating[session])
                        sessionnum[session] += int(sessions[session])
                    except:
                        None
            print(ratingnum, sessionnum)
            for session in ratingnum:
                ratingnum[session] = round(ratingnum[session]/sessionnum[session], 2)
            return ratingnum
        else:
            raise ValueError
    
    def siteratings(self, sitename, sitecolumn, sessioncolumn, ratingcolumn):
        sites, siterange = CH_Automation.columnrange(self, self.name, sitecolumn), []
        for i in sites:
            if i == sitename:
                siterange = sites[i]
        return self.totalsessionratings(sessioncolumn, ratingcolumn, siterange, sort = True)
    
    
    


c = Daily_Survey(r"C:\Users\jasoli3\Downloads\Cisco High Externship Daily Survey (1).xlsx")
#print(c.columnrange('Daily_Survey', 2))
#print(c.totalsessionratings([4, 7, 8], [5, 9, 10]))
#print(c.siteratings('ATL', 2, [4,7,8], [5,9,10]))
print(c.sessionratings(4,5,[0,60], True))

