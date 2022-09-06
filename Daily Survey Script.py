# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
workbook = load_workbook(filename=r"C:\Users\jasoli3\Downloads\Cisco High Externship Daily Survey (1).xlsx")
sheet = workbook.active

def uniquesites(initialrow, finalrow):
    uniqsite = {}
    columnnum = 1
    while sheet.cell(row = 1, column = columnnum).value != None:
        if sheet.cell(row = 1, column = columnnum).value == 'Cohort':
            break
        columnnum += 1
    i, j1, j2 = 1,initialrow,initialrow
    uniqsite[i] = [sheet.cell(row = initialrow, column = columnnum).value, 2]
    if finalrow == 'End':
        while sheet.cell(row = j1, column = columnnum).value != None:
            if sheet.cell(row = j1, column = columnnum).value != sheet.cell(row = j2, column = columnnum).value:
                uniqsite[i].append(j1-1)
                j2 = j1
                i += 1
                uniqsite[i] = [sheet.cell(row = j2, column = columnnum).value, j2]
            j1 += 1
        uniqsite[i].append(j1)
    elif type(finalrow) == int:
        while j1 < finalrow:
            if sheet.cell(row = j1, column = columnnum).value != None:
                if sheet.cell(row = j1, column = columnnum).value != sheet.cell(row = j2, column = columnnum).value:
                    uniqsite[i].append(j1-1)
                    j2 = j1
                    i += 1
                    uniqsite[i] = [sheet.cell(row = j2, column = columnnum).value, j2]
            j1 += 1
        uniqsite[i].append(j1)
    return uniqsite

def uniquevalcount(initialrow, finalrow, columnnum):
    uniqval = {}
    i, j = initialrow, initialrow
    list1 = []
    if finalrow == 'End':
        while sheet.cell(row = i, column = columnnum).value != None:
            list1.append(sheet.cell(row = i, column = columnnum).value)
            i += 1
        list1 = np.array(list1)
        list1 = np.unique(list1)
        for i in range(len(list1)):
            uniqval[list1[i]] = 0
        while sheet.cell(row = j, column = columnnum).value != None:
            for k in range(len(list1)):
                if sheet.cell(row = j, column = columnnum).value == list1[k]:
                    uniqval[list1[k]] += 1
            j += 1  
    elif type(finalrow) == int:
        while i < finalrow:
            if sheet.cell(row = i, column = columnnum).value != None:
                list1.append(sheet.cell(row = i, column = columnnum).value)
            i += 1
        list1 = np.array(list1)
        list1 = np.unique(list1)
        for i in range(len(list1)):
            uniqval[list1[i]] = 0
        while j <= finalrow:
            if sheet.cell(row = j, column = columnnum).value != None:
                for k in range(len(list1)):
                    if sheet.cell(row = j, column = columnnum).value == list1[k]:
                        uniqval[list1[k]] += 1
            j += 1
    return list1, uniqval

def sessionratings(initialrow, finalrow, sessioncolumn1, sessioncolumn2, sessioncolumn3, ratingcolumn1, ratingcolumn2, ratingcolumn3):
    i = initialrow
    session1, sessionscount1 = uniquevalcount(initialrow, finalrow, sessioncolumn1)
    session2, sessionscount2 = uniquevalcount(initialrow, finalrow, sessioncolumn2)
    session3, sessionscount3 = uniquevalcount(initialrow, finalrow, sessioncolumn3)
    session4, sessionscount = {}, {}
    totsessions = session1.tolist() + session2.tolist() + session3.tolist()
    totsessions = np.array(totsessions)
    totsessions = np.unique(totsessions)
    totsessions = totsessions.tolist()
    for j in range(len(totsessions)):
        session4[totsessions[j]] = 0
        sessionscount[totsessions[j]] = 0
    for value in sessionscount1:
        for value1 in sessionscount:
            if value == value1:
                sessionscount[value1] += sessionscount1[value]
    for value in sessionscount2:
        for value1 in sessionscount:
            if value == value1:
                sessionscount[value1] += sessionscount2[value]
    for value in sessionscount3:
        for value1 in sessionscount:
            if value == value1:
                sessionscount[value1] += sessionscount3[value]
    while i <= finalrow:
        if sheet.cell(row = i, column = sessioncolumn1).value != None and sheet.cell(row = i, column = ratingcolumn1).value != None:
            for k in range(len(totsessions)):                
                if sheet.cell(row = i, column = sessioncolumn1).value == totsessions[k]:
                    session4[totsessions[k]] += int(str(sheet.cell(row = i, column = ratingcolumn1).value)[0])
        if sheet.cell(row = i, column = sessioncolumn2).value != None and sheet.cell(row = i, column = ratingcolumn2).value != None:
            for k in range(len(totsessions)):                
                if sheet.cell(row = i, column = sessioncolumn2).value == totsessions[k]:
                    session4[totsessions[k]] += int(str(sheet.cell(row = i, column = ratingcolumn2).value)[0])
        if sheet.cell(row = i, column = sessioncolumn2).value != None and sheet.cell(row = i, column = ratingcolumn3).value != None:
            for k in range(len(totsessions)):                
                if sheet.cell(row = i, column = sessioncolumn3).value == totsessions[k]:
                    session4[totsessions[k]] += int(str(sheet.cell(row = i, column = ratingcolumn3).value)[0])
        i += 1
    for value in session4:
        session4[value] = round(int(session4[value])/int(sessionscount[value]), 2)
    return session4
        
def siteratings(initialrow, site, sessioncolumn1, sessioncolumn2, sessioncolumn3, ratingcolumn1, ratingcolumn2, ratingcolumn3):
    sitelist = uniquesites(initialrow, 'End')
    for i in range(1, len(sitelist)):
        if sitelist[i][0] == site:
            siterang = sitelist[i][1:3]
    stat = sessionratings(siterang[0], siterang[1], sessioncolumn1, sessioncolumn2, sessioncolumn3, ratingcolumn1, ratingcolumn2, ratingcolumn3)
    return stat

def bargraph(initialrow, finalrow, sessioncolumn1, sessioncolumn2, sessioncolumn3, ratingcolumn1, ratingcolumn2, ratingcolumn3):
    data = sessionratings(initialrow, finalrow, sessioncolumn1, sessioncolumn2, sessioncolumn3, ratingcolumn1, ratingcolumn2, ratingcolumn3)
    sessions = list(data.keys())
    ratings = list(data.values())
    plt.bar(sessions, ratings)
    plt.title('Session Ratings')

def sitebargraph(initialrow, site, sessioncolumn1, sessioncolumn2, sessioncolumn3, ratingcolumn1, ratingcolumn2, ratingcolumn3):
    data = siteratings(initialrow, site, sessioncolumn1, sessioncolumn2, sessioncolumn3, ratingcolumn1, ratingcolumn2, ratingcolumn3)
    sessions = list(data.keys())
    ratings = list(data.values())
    plt.bar(sessions, ratings)
    plt.title('Site Session Ratings')
    
def externattendance(initialrow, namecolumn, sessioncolumn, site, externname):
    sitelist = uniquesites(initialrow, 'End')
    for i in range(1, len(sitelist)):
        if sitelist[i][0] == site:
            siterang = sitelist[i][1:3]
    sessionlist, sessionval = uniquevalcount(siterang[0], siterang[1], sessioncolumn)
    sessionlist = sessionlist.tolist()
    for i in range(siterang[0], siterang[1]+1):
        if sheet.cell(row = i, column = namecolumn).value == externname:
            try:
                sessionlist.remove(str(sheet.cell(row=i, column = sessioncolumn).value))
            except ValueError:
                None
    return sessionlist

def externattedancetotal(initialrow, namecolumn, sessioncolumn1, sessioncolumn2, sessioncolumn3, site, externname):
    sessionlist1 = externattendance(initialrow, namecolumn, sessioncolumn1, site, externname)
    sessionlist2 = externattendance(initialrow, namecolumn, sessioncolumn2, site, externname)
    sessionlist3 = externattendance(initialrow, namecolumn, sessioncolumn3, site, externname)
    sessionlisttot = sessionlist1 + sessionlist2 + sessionlist3
    sessionlisttot = np.array(sessionlisttot)
    sessionlisttot = np.unique(sessionlisttot)
    return sessionlisttot

print(sessionratings(2, 1436, 4, 7, 8, 5, 9, 10))