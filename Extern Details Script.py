# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

class Extern_Details(object):
    
    def __init__(self):
        self.uniquesites = {}
    
    def addXLSX(self, xlsx):
        '''Created a workbook to manipulate the xlsx file; has to be fill address in this format:''' #var.addXLSX(r"C:\Users\jasoli3\Downloads\2022 Extern Details.xlsx")
        workbook = load_workbook(filename=xlsx)
        self.sheet = workbook.active
        self.df = pd.read_excel(xlsx)

    def sort(self, columnnum, namecolumn = None):
        columnname = self.sheet.cell(row = 1, column = columnnum).value
        if namecolumn == None:
            self.df.sort_values(by=columnname, inplace = True)
        if type(namecolumn) == int:
            columnname1 = self.sheet.cell(row = 1, column = namecolumn).value
            self.df.sort_values(by=[columnname, columnname1], inplace = True)
        self.df.dropna(how = 'all', inplace = True)
        self.df.reset_index(drop = True, inplace = True)
        return self.df

    def uniquesite(self, sitecolumn, namecolumn = None):
        '''uniquesites works if the Excel self.sheet is sorted by site and the column of site is names 'Sites'.//
        initialrow and finalrow are type int for the initial and final roself.sheet to search through. The function returns//
        a dict of the unique sites in the Excel self.sheet and the range of the roself.sheet it spans.'''
        self.uniquesites = {}
        self.sort(sitecolumn, namecolumn)
        for i in self.df.index:
            if i == 0:
                self.uniquesites[self.df.values[i][sitecolumn-1]] = [i]
            else:
                if i < self.df.index[len(self.df.index)-1]:
                    if self.df.values[i][sitecolumn-1] != self.df.values[i+1][sitecolumn-1]:
                        self.uniquesites[self.df.values[i][sitecolumn-1]].append(i)
                        self.uniquesites[self.df.values[i+1][sitecolumn-1]] = [i+1]
                else:
                    self.uniquesites[self.df.values[i][sitecolumn-1]].append(i)
        return self.uniquesites

    def uniquevalcount(self, initialrow, finalrow, columnnum):
        '''uniquevalcount returns a list and a dict, the list is of all of the unique values in a given column//
        the dict gives the number of occurances of each unique values. initialrow, finalrow, and columnnum are all ints//
        initialrow and finalrow are the roself.sheet the function searches through, and the columnnum is the'''
        uniqval = {}
        i, j = initialrow, initialrow
        list1 = []
        if finalrow == 'End':
            while self.sheet.cell(row = j, column = columnnum).value != None:
                list1.append(self.sheet.cell(row = i, column = columnnum).value)
                i += 1
            list1 = np.array(list1)
            list1 = np.unique(list1)
            for i in range(len(list1)):
                uniqval[list1[i]] = 0
            while self.sheet.cell(row = j, column = columnnum).value != None:
                for k in range(len(list1)):
                    if self.sheet.cell(row = j, column = columnnum).value == list1[k]:
                        uniqval[list1[k]] += 1
                j += 1  
        elif type(finalrow) == int:
            while i < finalrow:
                if self.sheet.cell(row = i, column = columnnum).value != None:
                    list1.append(self.sheet.cell(row = i, column = columnnum).value)
                i += 1
            list1 = np.array(list1)
            list1 = np.unique(list1)
            for i in range(len(list1)):
                uniqval[list1[i]] = 0
            while j <= finalrow:
                if self.sheet.cell(row = j, column = columnnum).value != None:
                    for k in range(len(list1)):
                        if self.sheet.cell(row = j, column = columnnum).value == list1[k]:
                            uniqval[list1[k]] += 1
                j += 1
        return list1, uniqval

    def demostatistics(self, initialrow, finalrow, columnnum):
        demostat = {}
        i = initialrow
        tot = 0
        if finalrow == 'End':
            while self.sheet.cell(row = i, column = columnnum).value != None:
                i += 1
        elif type(finalrow) == int:
            while i <= finalrow:
                if self.sheet.cell(row = i, column = columnnum).value != None:
                    tot += 1
                i += 1
        list1, uniqval = self.uniquevalcount(initialrow, finalrow, columnnum)
        for k in range(len(uniqval)):
            demostat[list1[k]] = str(round(uniqval[list1[k]]/tot * 100, 2)) + '%'
        return demostat

    def sitestat(self, initialrow, finalrow, columnnum, site):
        sitelist = self.uniquesites(initialrow, finalrow)
        for i in range(1, len(sitelist)):
            if sitelist[i][0] == site:
                siterang = sitelist[i][1:3]
        stat1, stat2 = self.uniquevalcount(siterang[0], siterang[1], columnnum)[1], self.demostatistics(siterang[0], siterang[1], columnnum)
        return stat1, stat2

    def piechart(self, initialrow, finalrow, columnnum):
        title = self.sheet.cell(row = 1, column = columnnum).value + ' Ratio'
        list1, uniqval = self.uniquevalcount(initialrow, finalrow, columnnum)
        values = []
        for i in uniqval:
            values.append(uniqval[i])
        values = np.array(values)
        plt.pie(values, labels = list1)
        plt.title(title)
        plt.show()

    def sitepiechart(self, initialrow, finalrow, columnnum, site):
        sitelist = self.uniquesites(initialrow, finalrow)
        for i in range(1, len(sitelist)):
            if sitelist[i][0] == site:
                siterang = sitelist[i][1:3]
        self.piechart(siterang[0], siterang[1], columnnum)

c = Extern_Details()
c.addXLSX(r"C:\Users\jasoli3\Downloads\2022 Extern Details.xlsx")
#print(c.sort(2, 1))
print(c.uniquesite(2,1))