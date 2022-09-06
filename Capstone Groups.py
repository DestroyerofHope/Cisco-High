# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import numpy as np


class Capstone_Groups():
    '''This is written as a class to have consistent global variables that can be manipulated through all functions.'''

    def __init__(self):
        '''This defines the class, var = Capstone_Groups()'''
        self.attr = {}
        self.groups = []
        self.groupnum = 0
        self.groupsize = []
        self.groupavg = []
        self.placegroup = []
        self.xlsx = None
        self.sheet = None

    def addXLSX(self, xlsx):
        '''Created a workbook to manipulate the xlsx file; has to be fill address in this format:''' #var.addXLSX(r"C:\Users\jasoli3\Downloads\2022 Extern Details.xlsx")
        workbook = load_workbook(filename=xlsx)
        self.sheet = workbook.active

    def assignNum(self, attribute):
        '''This function assigns the numerical value associate with a certian demographic attribute. attribute is a str. Outputs an int'''
        if attribute == 'RTP':
            return 0
        elif attribute == 'Richardson':
            return .33
        elif attribute == 'Herndon':
            return .67
        elif attribute == 'California':
            return 1
        elif attribute == 'Atlanta':
            return 0
        elif attribute == 'Chicago':
            return .25
        elif attribute == 'NYC':
            return .5
        elif attribute == 'Toronto':
            return .75
        elif attribute == 'St. Louis':
            return 1
        elif attribute == 'Male' or attribute == 'Man':
            return 0
        elif attribute == 'Female' or attribute == 'Woman':
            return 1
        elif attribute == 'Transgender':
            return 1.5
        elif attribute in ('African American', 'Black', 'Black / African American', '/African American'):
            return 0
        elif attribute == 'Latin / Spanish' or attribute == 'Spanish / Hispanic / Latino':
            return .17
        elif attribute == 'Pacific Islander':
            return .33
        elif attribute == 'Caucasian' or attribute == 'White / Caucasian':
            return .5
        elif attribute == 'Asian':
            return .67
        elif attribute == 'Other':
            return .83
        elif attribute == 'Prefer Not to Answer':
            return 1
        elif attribute == None:
            return .5

    def defAttributes(self, initialrow, finalrow, namecolumn, sitecolumn, gendercolumn, racecolumn):
        '''defAttribute creates a dict with the names of the externs and the numerical value associated with them. initialrow & finalrow are ints, representing the initial and final
            rows the algorithm searches through, namecolumn, sitecolumn, gendercolumn, and racecolumn are ints, they are the columns those details are located in.'''
        for i in range(initialrow, finalrow + 1):
            if self.sheet.cell(row=i, column=namecolumn).value != None:
                self.attr[self.sheet.cell(row=i, column=namecolumn).value] = [self.assignNum(self.sheet.cell(row=i, column=sitecolumn).value), self.assignNum(self.sheet.cell(
                    row=i, column=gendercolumn).value), self.assignNum(self.sheet.cell(row=i, column=racecolumn).value)]  # , self.sheet.cell(row = i, column = techcolumn).value

    def groupSizes(self, groupnum):
        '''groupSizes splits the total number of externs into equal groups, groupnum is an int representing the number of groups the externs are to be split into.
            Outputs a list of group sizes'''
        self.groupnum = groupnum
        num, remainder = np.floor(
            len(self.attr)/self.groupnum), (len(self.attr) % self.groupnum)
        for i in range(self.groupnum):
            self.groupsize.append(num)
        for i in range(remainder):
            self.groupsize[i] += 1
        for i in range(len(self.groupsize)):
            self.groupsize[i] = int(self.groupsize[i])
        return self.groupsize

    def assignTarget(self):
        '''assignTarget calculates the target numbers for different demographic attributes based on the demographic ratios in the sample. Outputs a list of target values.'''
        targets = []
        for i in range(len(list(self.attr.values())[0])):
            target = 0
            for j in range(len(self.attr)):
                target += list(self.attr.values())[j][i]
            target = target/int(len(self.attr))
            targets.append(target)
        return targets

    def createGroup(self, groups, iterations, groupiter):
        '''createGroup recursively creates a group based on the numerical values calculated in assignTarget. iterations and groupiter are ints, groups is a dict
            iterations is the number of iterations the function has done already and groupiter is the number of people in the created group, 
            groups is the dict of externs and demographic values. Outputs a list of capstone groups and the experimental demographic numerical values for the groups.'''
        targets = self.assignTarget()
        if iterations == 0:
            self.placegroup.append(
                [list(groups.keys())[0], list(groups.values())[0]])
            del groups[list(groups.keys())[0]]
            self.groupavg.append(
                [list(self.placegroup)[0][1][0], len(self.placegroup)])
            self.groupavg.append(
                [list(self.placegroup)[0][1][1], len(self.placegroup)])
            self.groupavg.append(
                [list(self.placegroup)[0][1][2], len(self.placegroup)])
        avgnum = {}
        for name in groups:
            avgs = [(list(groups[name])[0] + (self.groupavg[0][0]*self.groupavg[0][1]))/(self.groupavg[0][1]+1), (list(groups[name])[1] + (self.groupavg[1]
                    [0]*self.groupavg[1][1]))/(self.groupavg[1][1]+1), (list(groups[name])[2] + (self.groupavg[2][0]*self.groupavg[2][1]))/(self.groupavg[2][1]+1)]
            for j in range(len(targets)):
                avgs[j] = abs(targets[j] - avgs[j])
            placeavgnum = sum(avgs)/3
            avgnum[name] = placeavgnum
        choice = list(avgnum.keys())[0]
        for name in avgnum:
            if avgnum[name] < avgnum[choice]:
                choice = name
        self.placegroup.append([choice, groups[choice]])
        self.groupavg[0][0] = self.groupavg[0][0] * \
            self.groupavg[0][1] + groups[choice][0]
        self.groupavg[1][0] = self.groupavg[1][0] * \
            self.groupavg[1][1] + groups[choice][1]
        self.groupavg[2][0] = self.groupavg[2][0] * \
            self.groupavg[2][1] + groups[choice][2]
        self.groupavg[0][1] += 1
        self.groupavg[1][1] += 1
        self.groupavg[2][1] += 1
        self.groupavg[0][0] = self.groupavg[0][0]/self.groupavg[0][1]
        self.groupavg[1][0] = self.groupavg[1][0]/self.groupavg[1][1]
        self.groupavg[2][0] = self.groupavg[2][0]/self.groupavg[2][1]
        groups.pop(choice)
        choice = ''
        if iterations < groupiter - 2:
            self.createGroup(groups, iterations + 1, groupiter)
        if iterations <= groupiter - 2:
            del groups
            return self.placegroup, self.groupavg

    def assignGroups(self):
        '''Calls createGroup multiple times to create the appropriate number of groups for the externs. It returns a list of the groups.'''
        groupiter, self.placegroup, grouptarget = self.groupsize, [], []
        groups = self.attr.copy()
        for size in groupiter:
            self.groupavg, self.placegroup = [], []
            groupnames = []
            group, groupavg = self.createGroup(groups, 0, size)
            grouptarget.append([groupavg[0][0], groupavg[1][0], groupavg[2][0]])
            for i in range(len(group)):
                groupnames.append(group[i][0])
            self.groups.append(groupnames)
        return self.groups, grouptarget
    
    def targetError(self):
        '''targetError calculates the percent error of the groups made by assignGroups in comparison to the calculated target numbers in assingTarget
            It returns a list of the average percent errors of each group.'''
        target = self.assignTarget()
        groups, grouptarget = self.assignGroups()
        error = []
        for i in range(len(grouptarget)):
            error1 = abs(grouptarget[i][0] - target[0])/ target[0]
            error2 = abs(grouptarget[i][1] - target[1])/ target[1]
            error3 = abs(grouptarget[i][2] - target[2])/ target[2]
            error.append(str(round(((error1 + error2 + error3)/3)*100, 2)) + ' %')
        return error

    def __str__(self):
        '''If groups are created, printing the class returns the groups, if not, it returns the attributes dict.'''
        if len(self.groups) != 0:
            return str(self.groups)
        else:
            return str(self.attr)


c = Capstone_Groups()
c.addXLSX(r"C:\Users\jasoli3\Downloads\2022 Extern Details.xlsx")
c.defAttributes(2, 87, 1, 2, 4, 5)
c.groupSizes(8)
print(c.assignTarget())
print(c.assignGroups())
print(c.targetError())
