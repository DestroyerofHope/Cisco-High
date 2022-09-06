# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
workbook = load_workbook(filename=r"C:\Users\jasoli3\Downloads\2022 Capstone Rubric (2).xlsx")
sheet = workbook.active

def uniqueteams(initialrow, finalrow):
    uniqteam = {}
    columnnum = 1
    while sheet.cell(row = 1, column = columnnum).value != None:
        if sheet.cell(row = 1, column = columnnum).value == 'Group Name':
            break
        columnnum += 1
    i, j1, j2 = 1,initialrow,initialrow
    uniqteam[i] = [sheet.cell(row = initialrow, column = columnnum).value, 2]
    if finalrow == 'End':
        while sheet.cell(row = j1, column = columnnum).value != None:
            if sheet.cell(row = j1, column = columnnum).value != sheet.cell(row = j2, column = columnnum).value:
                uniqteam[i].append(j1-1)
                j2 = j1
                i += 1
                uniqteam[i] = [sheet.cell(row = j2, column = columnnum).value, j2]
            j1 += 1
        uniqteam[i].append(j1)
    elif type(finalrow) == int:
        while j1 < finalrow:
            if sheet.cell(row = j1, column = columnnum).value != None:
                if sheet.cell(row = j1, column = columnnum).value != sheet.cell(row = j2, column = columnnum).value:
                    uniqteam[i].append(j1-1)
                    j2 = j1
                    i += 1
                    uniqteam[i] = [sheet.cell(row = j2, column = columnnum).value, j2]
            j1 += 1
        uniqteam[i].append(j1)
    return uniqteam

def totalscore(initialrow, finalrow, initialcolumn, finalcolumn, teamname):
    teamslist, teamrang, score, tot = uniqueteams(initialrow, finalrow), {}, {}, 0
    for i in range(1, len(teamslist)+1):
        if teamslist[i][0] == teamname:
            teamrang = teamslist[i][1:3]
    for i in range(teamrang[0], teamrang[1]+1):
        score[i] = 0
    for i in range(teamrang[0], teamrang[1]+1):
        for j in range(initialcolumn, finalcolumn+2):
            if sheet.cell(row = i, column = j).value != None:
                score[i] += int(str(sheet.cell(row = i, column = j).value)[0])
    for i in score:
        tot += score[i]
    tot = tot/len(score)
    return tot

def lowhighscores(initialrow, finalrow, initialcolumn, finalcolumn, teamname):
    teamslist, teamrang, score, high, low = uniqueteams(initialrow, finalrow), {}, {}, [], []
    for i in range(1, len(teamslist)):
        if teamslist[i][0] == teamname:
            teamrang = teamslist[i][1:3]
    for i in range(initialcolumn, finalcolumn+1):
        score[sheet.cell(row=1, column=i).value] = 0
    for i in range(initialcolumn, finalcolumn+1):
        for j in range(teamrang[0], teamrang[1]+1):
            if sheet.cell(row=j, column=i).value != None:
                score[sheet.cell(row=1, column=i).value] += int(str(sheet.cell(row=j, column=i).value)[0])
    high.append([list(score.keys())[0], list(score.values())[0]])
    low.append([list(score.keys())[0], list(score.values())[0]])
    for i in range(1, len(score)):
        if list(score.values())[i] > list(high[0])[1]:
            high.clear()
            high.append([list(score.keys())[i], list(score.values())[i]])
        elif list(score.values())[i] == list(high[0])[1]:
            high.append([list(score.keys())[i], list(score.values())[i]])
        if list(score.values())[i] < list(low[0])[1]:
            low.clear()
            low.append([list(score.keys())[i], list(score.values())[i]])
        elif list(score.values())[i] == list(low[0])[1]:
            low.append([list(score.keys())[i], list(score.values())[i]]) 
    return 'Highest: ' + str(high), 'Lowest: ' + str(low)

def winners(initialrow, finalrow, initialcolumn, finalcolumn):
    teamlist, teamscores = uniqueteams(initialrow, finalrow), {}
    for i in range(len(teamlist)):
        teamscores[list(teamlist.values())[i][0]] = totalscore(initialrow, finalrow, initialcolumn, finalcolumn, list(teamlist.values())[i][0])   
    teamscores = dict(sorted(teamscores.items(), key=lambda item: item[1]))
    return teamscores
    
print(winners(2,54,2,9))